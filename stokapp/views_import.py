from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db import transaction
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import StokItem, Tedarikci, Musteri, Kategori


@login_required
def veri_ice_aktarma(request):
    """Veri içe aktarma ana sayfası"""
    return render(request, 'stokapp/veri_ice_aktarma.html')


@login_required
def download_stok_template(request):
    """Stok içe aktarma şablonu indir"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Şablonu"
    
    # Başlık satırı
    headers = [
        'Stok Kodu', 'Ürün Adı', 'Kategori', 'Birim', 'Barkod',
        'Mevcut Miktar', 'Minimum Stok', 'Alış Fiyatı', 'Para Birimi',
        'Açıklama', 'Tedarikçi'
    ]
    
    # Başlıkları yaz ve stil ver
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Örnek veri satırları
    example_data = [
        ['STOK-001', 'Örnek Ürün 1', 'Ham Madde', 'Adet', '1234567890123', 
         100, 10, 25.50, 'TL', 'Örnek açıklama', 'Tedarikçi 1'],
        ['STOK-002', 'Örnek Ürün 2', 'Yarı Mamül', 'Kg', '9876543210987',
         50, 5, 15.75, 'TL', 'Başka bir açıklama', 'Tedarikçi 2'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Sütun genişliklerini ayarla
    column_widths = [15, 30, 15, 10, 15, 12, 12, 12, 12, 30, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stok_import_sablonu.xlsx"'
    
    wb.save(response)
    return response


@login_required
def download_tedarikci_template(request):
    """Tedarikçi içe aktarma şablonu indir"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tedarikçi Şablonu"
    
    headers = ['Ad', 'Telefon', 'E-posta', 'Adres']
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    example_data = [
        ['Tedarikçi 1', '0212 123 45 67', 'tedarikci1@example.com', 'İstanbul, Türkiye'],
        ['Tedarikçi 2', '0312 987 65 43', 'tedarikci2@example.com', 'Ankara, Türkiye'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    column_widths = [30, 15, 30, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tedarikci_import_sablonu.xlsx"'
    
    wb.save(response)
    return response


@login_required
def download_musteri_template(request):
    """Müşteri içe aktarma şablonu indir"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Müşteri Şablonu"
    
    headers = ['Ad', 'Telefon', 'E-posta', 'Adres']
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    example_data = [
        ['Müşteri 1', '0212 111 22 33', 'musteri1@example.com', 'İstanbul, Türkiye'],
        ['Müşteri 2', '0312 444 55 66', 'musteri2@example.com', 'Ankara, Türkiye'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    column_widths = [30, 15, 30, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="musteri_import_sablonu.xlsx"'
    
    wb.save(response)
    return response


@login_required
def import_stok_excel(request):
    """Stok Excel dosyasını içe aktar"""
    if request.method == 'POST' and request.FILES.get('stok_file'):
        try:
            excel_file = request.FILES['stok_file']
            df = pd.read_excel(excel_file)
            
            # Gerekli sütunları kontrol et
            required_columns = ['Stok Kodu', 'Ürün Adı', 'Kategori', 'Birim']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Eksik sütunlar: {", ".join(missing_columns)}')
                return redirect('stokapp:veri_ice_aktarma')
            
            success_count = 0
            error_count = 0
            errors = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Kategoriyi bul veya oluştur
                        kategori_ad = str(row['Kategori']).strip()
                        kategori, created = Kategori.objects.get_or_create(
                            ad=kategori_ad,
                            defaults={'stok_tipi': 'HAM_MADDE'}
                        )
                        
                        # Birim alanı (StokItem'da CharField olarak tanımlı)
                        birim_ad = str(row.get('Birim', 'Adet')).strip()
                        
                        # Tedarikçiyi bul (varsa)
                        tedarikci = None
                        if 'Tedarikçi' in row and pd.notna(row['Tedarikçi']):
                            tedarikci_ad = str(row['Tedarikçi']).strip()
                            tedarikci, created = Tedarikci.objects.get_or_create(ad=tedarikci_ad)
                        
                        # Stok kodunu kontrol et
                        stok_kodu = str(row['Stok Kodu']).strip()
                        if not stok_kodu:
                            error_count += 1
                            errors.append(f'Satır {index + 2}: Stok kodu boş olamaz')
                            continue
                        
                        # StokItem oluştur veya güncelle
                        stok_item, created = StokItem.objects.update_or_create(
                            stok_kodu=stok_kodu,
                            defaults={
                                'ad': str(row['Ürün Adı']).strip(),
                                'kategori': kategori,
                                'birim': birim_ad,
                                'mevcut_miktar': float(row.get('Mevcut Miktar', 0)) if pd.notna(row.get('Mevcut Miktar')) else 0,
                                'minimum_stok': float(row.get('Minimum Stok', 0)) if pd.notna(row.get('Minimum Stok')) else 0,
                                'alis_fiyati': float(row.get('Alış Fiyatı', 0)) if pd.notna(row.get('Alış Fiyatı')) else 0,
                                'alis_para_birimi': str(row.get('Para Birimi', 'TL')).strip(),
                                'barkod': str(row.get('Barkod', '')).strip(),
                                'aciklama': str(row.get('Açıklama', '')).strip(),
                                'tedarikci': tedarikci,
                            }
                        )
                        
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f'Satır {index + 2}: {str(e)}')
            
            if success_count > 0:
                messages.success(request, f'{success_count} stok başarıyla içe aktarıldı.')
            if error_count > 0:
                messages.warning(request, f'{error_count} kayıt içe aktarılamadı. İlk 5 hata: {", ".join(errors[:5])}')
            
            return redirect('stokapp:veri_ice_aktarma')
            
        except Exception as e:
            messages.error(request, f'İçe aktarma hatası: {str(e)}')
            return redirect('stokapp:veri_ice_aktarma')
    
    return redirect('stokapp:veri_ice_aktarma')


@login_required
def import_tedarikci_excel(request):
    """Tedarikçi Excel dosyasını içe aktar"""
    if request.method == 'POST' and request.FILES.get('tedarikci_file'):
        try:
            excel_file = request.FILES['tedarikci_file']
            df = pd.read_excel(excel_file)
            
            required_columns = ['Ad']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Eksik sütunlar: {", ".join(missing_columns)}')
                return redirect('stokapp:veri_ice_aktarma')
            
            success_count = 0
            error_count = 0
            errors = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        ad = str(row['Ad']).strip()
                        if not ad:
                            error_count += 1
                            errors.append(f'Satır {index + 2}: Ad boş olamaz')
                            continue
                        
                        tedarikci, created = Tedarikci.objects.update_or_create(
                            ad=ad,
                            defaults={
                                'telefon': str(row.get('Telefon', '')).strip() if pd.notna(row.get('Telefon')) else '',
                                'email': str(row.get('E-posta', '')).strip() if pd.notna(row.get('E-posta')) else '',
                                'adres': str(row.get('Adres', '')).strip() if pd.notna(row.get('Adres')) else '',
                            }
                        )
                        
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f'Satır {index + 2}: {str(e)}')
            
            if success_count > 0:
                messages.success(request, f'{success_count} tedarikçi başarıyla içe aktarıldı.')
            if error_count > 0:
                messages.warning(request, f'{error_count} kayıt içe aktarılamadı. İlk 5 hata: {", ".join(errors[:5])}')
            
            return redirect('stokapp:veri_ice_aktarma')
            
        except Exception as e:
            messages.error(request, f'İçe aktarma hatası: {str(e)}')
            return redirect('stokapp:veri_ice_aktarma')
    
    return redirect('stokapp:veri_ice_aktarma')


@login_required
def import_musteri_excel(request):
    """Müşteri Excel dosyasını içe aktar"""
    if request.method == 'POST' and request.FILES.get('musteri_file'):
        try:
            excel_file = request.FILES['musteri_file']
            df = pd.read_excel(excel_file)
            
            required_columns = ['Ad']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Eksik sütunlar: {", ".join(missing_columns)}')
                return redirect('stokapp:veri_ice_aktarma')
            
            success_count = 0
            error_count = 0
            errors = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        ad = str(row['Ad']).strip()
                        if not ad:
                            error_count += 1
                            errors.append(f'Satır {index + 2}: Ad boş olamaz')
                            continue
                        
                        musteri, created = Musteri.objects.update_or_create(
                            ad=ad,
                            defaults={
                                'telefon': str(row.get('Telefon', '')).strip() if pd.notna(row.get('Telefon')) else '',
                                'email': str(row.get('E-posta', '')).strip() if pd.notna(row.get('E-posta')) else '',
                                'adres': str(row.get('Adres', '')).strip() if pd.notna(row.get('Adres')) else '',
                            }
                        )
                        
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f'Satır {index + 2}: {str(e)}')
            
            if success_count > 0:
                messages.success(request, f'{success_count} müşteri başarıyla içe aktarıldı.')
            if error_count > 0:
                messages.warning(request, f'{error_count} kayıt içe aktarılamadı. İlk 5 hata: {", ".join(errors[:5])}')
            
            return redirect('stokapp:veri_ice_aktarma')
            
        except Exception as e:
            messages.error(request, f'İçe aktarma hatası: {str(e)}')
            return redirect('stokapp:veri_ice_aktarma')
    
    return redirect('stokapp:veri_ice_aktarma')

